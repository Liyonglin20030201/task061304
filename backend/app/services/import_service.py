import pandas as pd
import numpy as np
from datetime import datetime, timezone
from typing import Dict, List
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from sqlalchemy import text
from app.config import get_settings
from app.services.cleaning_service import clean_dataframe

settings = get_settings()

import_progress: Dict[int, dict] = {}


async def process_import_file(task_id: int, file_path: str, data_type: str):
    engine = create_async_engine(settings.DATABASE_URL)
    async_session = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    import_progress[task_id] = {"status": "processing", "progress": 0, "processed_rows": 0, "total_rows": 0}

    # Use a separate session for status updates (these commit independently)
    async with async_session() as status_db:
        await status_db.execute(
            text("UPDATE import_tasks SET status='processing', started_at=:now WHERE id=:id"),
            {"id": task_id, "now": datetime.now(timezone.utc)},
        )
        await status_db.commit()

    async with async_session() as db:
        try:
            if file_path.endswith(".csv"):
                total_rows = 0
                with open(file_path, encoding="utf-8", errors="ignore") as f:
                    for _ in f:
                        total_rows += 1
                total_rows = max(total_rows - 1, 0)
            else:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                total_rows = ws.max_row - 1 if ws.max_row else 0
                wb.close()

            import_progress[task_id]["total_rows"] = total_rows

            async with async_session() as status_db:
                await status_db.execute(
                    text("UPDATE import_tasks SET total_rows=:total WHERE id=:id"),
                    {"id": task_id, "total": total_rows},
                )
                await status_db.commit()

            chunk_size = settings.CHUNK_SIZE
            processed = 0
            success = 0
            errors = 0
            duplicates = 0

            if file_path.endswith(".csv"):
                reader = pd.read_csv(file_path, chunksize=chunk_size, encoding="utf-8", on_bad_lines="skip")
            else:
                df_full = pd.read_excel(file_path)
                reader = [df_full[i:i + chunk_size] for i in range(0, len(df_full), chunk_size)]

            # All inserts happen within a single transaction (no intermediate commits)
            # If any batch fails, the entire import is rolled back
            for chunk in reader:
                cleaned, chunk_errors, chunk_dupes = clean_dataframe(chunk, data_type)
                batch_success = await _insert_batch_no_commit(db, cleaned, data_type)
                processed += len(chunk)
                success += batch_success
                errors += chunk_errors
                duplicates += chunk_dupes

                progress_pct = round(processed / max(total_rows, 1) * 100, 1)
                import_progress[task_id].update({
                    "progress": progress_pct,
                    "processed_rows": processed,
                    "success_rows": success,
                    "error_rows": errors,
                    "duplicate_rows": duplicates,
                    "status": "processing",
                })

            # All batches succeeded — commit the entire transaction atomically
            await db.commit()

            async with async_session() as status_db:
                await status_db.execute(
                    text("""UPDATE import_tasks SET status='completed', completed_at=:now,
                            processed_rows=:p, success_rows=:s, error_rows=:e, duplicate_rows=:d
                            WHERE id=:id"""),
                    {"id": task_id, "now": datetime.now(timezone.utc),
                     "p": processed, "s": success, "e": errors, "d": duplicates},
                )
                await status_db.commit()
                await _log_task(status_db, task_id, "info",
                    f"导入完成(原子提交): 总{total_rows}行, 成功{success}, 错误{errors}, 重复{duplicates}")

            import_progress[task_id].update({"status": "completed", "progress": 100.0})

        except Exception as e:
            # Rollback ALL inserted data — return to pre-import state
            await db.rollback()

            async with async_session() as status_db:
                await status_db.execute(
                    text("UPDATE import_tasks SET status='failed', error_details=:err WHERE id=:id"),
                    {"id": task_id, "err": str(e)},
                )
                await status_db.commit()
                await _log_task(status_db, task_id, "error",
                    f"导入失败，已回滚全部数据: {str(e)}")

            import_progress[task_id].update({"status": "failed", "error": str(e)})

    await engine.dispose()


async def _insert_batch_no_commit(db: AsyncSession, df: pd.DataFrame, data_type: str) -> int:
    """Insert a batch of records WITHOUT committing. Caller manages the transaction."""
    if df.empty:
        return 0

    table_map = {
        "sales": "sales",
        "inventory": "inventory",
        "members": "members",
        "promotions": "promotions",
        "traffic": "traffic",
        "weather": "weather",
    }
    table = table_map.get(data_type)
    if not table:
        return 0

    records = df.where(pd.notnull(df), None).to_dict("records")
    columns = list(records[0].keys())
    col_str = ", ".join(columns)
    val_placeholders = ", ".join([f":{c}" for c in columns])

    conflict_keys = {
        "sales": "store_id, receipt_no, item_id",
        "inventory": "store_id, item_id, snapshot_date",
        "traffic": "store_id, traffic_date, hour",
        "weather": "city, weather_date",
    }

    if data_type in conflict_keys:
        sql = f"""INSERT INTO {table} ({col_str}) VALUES ({val_placeholders})
                  ON CONFLICT ({conflict_keys[data_type]}) DO NOTHING"""
    else:
        sql = f"INSERT INTO {table} ({col_str}) VALUES ({val_placeholders})"

    inserted = 0
    for record in records:
        await db.execute(text(sql), record)
        inserted += 1
    # No commit here — transaction stays open
    return inserted


async def _log_task(db: AsyncSession, task_id: int, level: str, message: str):
    await db.execute(
        text("INSERT INTO task_logs (task_id, level, message, created_at) VALUES (:tid, :lvl, :msg, :now)"),
        {"tid": task_id, "lvl": level, "msg": message, "now": datetime.now(timezone.utc)},
    )
    await db.commit()
